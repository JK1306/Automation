import tabula
import os
# import camelot
import openpyxl
import csv
from openpyxl import load_workbook
from datetime import date
import pandas as pd
import glob
from openpyxl import workbook
t_date = date.today()
import pdfplumber
import sys
import logging
import datetime
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

logging.basicConfig(filename='Drr_generator.log',
                    format='%(asctime)s %(message)s',
                    filemode='a',
                    level = logging.DEBUG)


import configparser
config = configparser.ConfigParser()
config.read('DRR.config.ini')

properties_mapping = {"Night Audit- Hawthorn Suites El Paso Airport":"opera","Night Audit- Holiday Inn Express Abilene":"opera",
"Night Audit- Candlewood Monhans":"opera","Night Audit- Candlewood Abilene":"opera","Night Audit- Holiday Inn Express Van Horn":"opera",
"Night Audit- Candlewood Beaumont":"opera","Night Audit- Holiday Inn Express Chandler":"opera","Night Audit- Staybridge Pecos":"opera",
"Night Audit- Holiday Inn Express San Angelo":"opera","Night Audit- Holiday Inn Express- Las Vegas":"opera","Night Audit- Holiday Inn Express Fort Worth":"opera",
"Night Audit- Tryp Hotel College Station":"opera","Night Audit- Holiday Inn DFW Bedford":"opera","Night Audit- Towneplace Suites Houston NW":"Fossee_type1",
"Night Audit- Towneplace Plano":"Fossee_type1"}




def trial_balance(path,out_path,balance,balance_yesterday):
    logging.info('RAPBot has started to enter the trial balance in DRR_OPERA')
    wb_obj_temp = load_workbook("DRR_Template.xlsx")
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column
    sheet_obj_temp['C67'].value = balance[0]
    sheet_obj_temp['C68'].value = balance[1]
    sheet_obj_temp['C69'].value = balance[2]
    sheet_obj_temp['B67'].value = balance_yesterday[0]
    sheet_obj_temp['B68'].value = balance_yesterday[1]
    sheet_obj_temp['B69'].value = balance_yesterday[2]
    try:
        if ',' in balance[0] and ',' in balance_yesterday[0]:
            guest = balance[0].split(',')
            guest_ledg = guest[0]+guest[1]
            ledger = balance_yesterday[0].split(',')
            gues_led = ledger[0]+ledger[1]
            guest_ledger = float(guest_ledg) - float(gues_led)
            sheet_obj_temp['B55'].value = guest_ledger
        else:
            guest_ledger = float(balance[0]) - float(balance_yesterday[0])
            sheet_obj_temp['B55'].value = guest_ledger
        if ',' in balance[1] and ',' in balance_yesterday[1]:
            guest = balance[1].split(',')
            guest_ledg = guest[0]+guest[1]
            ledger = balance_yesterday[1].split(',')
            gues_led = ledger[0]+ledger[1]
            city_ledger = float(guest_ledg) - float(gues_led)
            sheet_obj_temp['B56'].value = city_ledger
        else:
            city_ledger = float(balance[1]) - float(balance_yesterday[1])
            sheet_obj_temp['B56'].value = city_ledger
        if ',' in balance[2] and ',' in balance_yesterday[2]:
            guest = balance[2].split(',')
            guest_ledg = guest[0]+guest[1]
            ledger = balance_yesterday[2].split(',')
            gues_led = ledger[0]+ledger[1]
            advance_deposit = float(guest_ledg) - float(gues_led)
            sheet_obj_temp['B57'].value = advance_deposit
        else:
            advance_deposit = float(balance[2]) - float(balance_yesterday[2])
            sheet_obj_temp['B57'].value = advance_deposit
    except Exception as e:
        print(e)
        logging.info('Error in placing guest ledger,city ledger and advance deposit ---->', e)



    pdf = pdfplumber.open(path)
    for page in pdf.pages:
        page = page.extract_text()
        drr_data = page.split('\n')
        for data in drr_data:
            if 'ar ledger payments' in data.lower():
                sheet_obj_temp['B51'].value = data.split(' ')[-1]
            if 'accommodation' in data.lower() or 'accomodation' in data.lower() or 'room' in data.lower():
                if not 'adj' in data.lower():
                    print(data.split(' ')[-1])
                    sheet_obj_temp['B25'].value = data.split(' ')[-1]
                
            if 'state tax' in data.lower():
                if not 'adj' in data.lower():
                    print(data.split(' ')[-1])
                    sheet_obj_temp['B39'].value = data.split(' ')[-1]

            
            if 'city tax' in data.lower():
                # print(data,'city tax')
                if not 'adj' in data.lower():
                    print(data.split(' ')[-1])
                    sheet_obj_temp['B40'].value = data.split(' ')[-1]

            if 'venue tax' in data.lower():
                if not 'adj' in data.lower():
                    print(data.split(' ')[-1])
                    sheet_obj_temp['B41'].value = data.split(' ')[-1]

            if 'sales tax' in data.lower():
                print(data.split(' ')[-1])
                sheet_obj_temp['B42'].value = data.split(' ')[-1]

            if 'sundries' in data.lower() or "pantry" in data.lower() or "cupboard" in data.lower() or "suite shop" in data.lower() or "gift shop" in data.lower():
                print(data.split(' ')[-1])
                sheet_obj_temp['B29'].value = data.split(' ')[-1]

            if 'cash' in data.lower():
                if not 'cashiers' in data.lower():
                    print(data.split(' ')[-1])
                    sheet_obj_temp['B45'].value = data.split(' ')[-1]

            if 'american' in data.lower():
                print(data)
                sheet_obj_temp['B46'].value = data.split(' ')[-1]
            if 'master' in data.lower():
                print(data)
                sheet_obj_temp['B47'].value = data.split(' ')[-1]
            if 'visa' in data.lower():
                print(data)
                sheet_obj_temp['B48'].value = data.split(' ')[-1]
            
            if 'discover' in data.lower():
                print(data)
                sheet_obj_temp['B49'].value = data.split(' ')[-1]

            if 'pet' in data.lower():
                print(data.split(' ')[-1])
                sheet_obj_temp['B28'].value = data.split(' ')[-1]
            
            if 'gift' in data.lower():
                print(data.split(' ')[-1])
                sheet_obj_temp['B29'].value = data.split(' ')[-1]
            
            if 'misc' in data.lower():
                print(data.split(' ')[-1])
                sheet_obj_temp['B36'].value = data.split(' ')[-1]

            
            
    os.makedirs(f"{out_path}\\DRR_excel",exist_ok=True)
    wb_obj_temp.save(f"{out_path}\\DRR_excel\\opera_drr.xlsx")
    pdf.close()
    logging.info('RAPBot has sucessfully entered the values in DRR')

def ledger_entry(path,out_path):
    balance = []
    balance_yesterday = []
    pdf = pdfplumber.open(path)
    for page in pdf.pages:
        page = page.extract_text()
        drr_data = page.split('\n')
        for data in drr_data:
            if 'balance today' in data.lower():
                balance.append(data.split(' ')[-1])
            if 'balance yesterday' in data.lower():
                balance_yesterday.append(data.split(' ')[-1])

    pdf.close()
    trial_balance(path,out_path,balance,balance_yesterday)

def drr_excel(oper_file_path):
    file_loc = config["file_path"]["loc"]
    y = datetime.datetime.now()
    mon = y.strftime("%B")
    t_date = date.today()
    yesterday = t_date-datetime.timedelta(days=1)
    previous_date = yesterday.strftime("%d")
    # for folder in os.listdir(f'{file_loc}'):
    folder = oper_file_path
    print(folder)
    # if 1>0:
    try:
        if folder in properties_mapping:
            for file_name in os.listdir(f'{file_loc}\\{folder}\\{mon}\\{previous_date}'):
                if 'market' in file_name.lower() or 'mcs' in file_name.lower():
                    df = tabula.read_pdf(f"{file_loc}\\{folder}\\{mon}\\{previous_date}\\{file_name}", pages = 'all')
                    page = len(df)
                    print(page)
                    excel_output_path = f"{file_loc}\\{folder}\\{mon}\\{previous_date}\\tabula_output\\market_code_stats_report"
                    name = 'market_code_stats_report'
                    for list_of_pages in range(page):
                        os.makedirs(excel_output_path, exist_ok=True)
                        df[list_of_pages].to_excel(f'{excel_output_path}\\{name}_{list_of_pages}.xlsx')

                if 'manager' in file_name.lower() or 'mf' in file_name.lower():
                    df = tabula.read_pdf(f"{file_loc}\\{folder}\\{mon}\\{previous_date}\\{file_name}", multiple_tables=True)
                    page = len(df)
                    out_path = f"{file_loc}\\{folder}\\{mon}\\{previous_date}\\tabula_output\\manager_report"
                    for list_of_pages in range(page):
                        os.makedirs(out_path, exist_ok=True)
                        df[list_of_pages].to_excel(f'{out_path}\\manager_report.xlsx')
                if 'trial' in file_name.lower() or 'tb' in file_name.lower():
                    ledger_entry(f'{file_loc}\\{folder}\\{mon}\\{previous_date}\\{file_name}',f'{file_loc}\\{folder}\\{mon}\\{previous_date}')
                    # trial_balance(f'{file_loc}\\{folder}\\{mon}\\{previous_date}\\{file_name}',f'{file_loc}\\{folder}\\{mon}\\{previous_date}')
            if page > 1:
                convert_tabula_to_drr(f'{file_loc}\\{folder}\\{mon}\\{previous_date}')
            elif page == 1:
                market_segmentation(f'{file_loc}\\{folder}\\{mon}\\{previous_date}')
        else:
            logging.info(f'{folder} property is not opera')
    except Exception as e:
        print(f'{folder} Not opera PMS system')
        print(e,'0000000')
        logging.info(f"{folder} is not Opera PMS System")
            

def market_segmentation(path):
    tabula_path=f'{path}\\tabula_output\\market_code_stats_report'
    wb_obj = load_workbook(f'{tabula_path}\\market_code_stats_report_0.xlsx')
    sheet_obj = wb_obj.active
    rows = sheet_obj.max_row
    cols = sheet_obj.max_column
    key = []
    val_room = []
    val_revenue = []
    (start_row,start_col)=marketformat(f"{tabula_path}\\market_code_stats_report_0.xlsx")
    for i in range(1,rows+1):
        cell_obj = sheet_obj.cell(row=i, column=start_col)
        keys = cell_obj.value
        key.append(keys.split(' ')[0])
    for i in range(1,rows+1):
        cell_obj = sheet_obj.cell(row=i, column=start_col+2)
        val = cell_obj.value
        val_room.append(val)

    for i in range(1,rows+1):
        cell_obj = sheet_obj.cell(row=i, column=start_col+3)
        val_rev = cell_obj.value
        val_revenue.append(val_rev)
       

    room = dict(zip(key, val_room))
    revenue = dict(zip(key,val_revenue))
    print (room)
    room = dict(zip(key, val_room))
    revenue = dict(zip(key,val_revenue))

    room_val(room,path)
    rev_val(revenue,path)
    manager_report(f'{path}\\tabula_output\\manager_report',path)



def room_val(room,output_path):
    
    wb_obj_temp = load_workbook(f"{output_path}\\DRR_excel\\opera_drr.xlsx")
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column
    try:
        sheet_obj_temp['B5'].value = room['A']
        print(room['A'])
    except:
        sheet_obj_temp['B5'].value = 0
    try:
        sheet_obj_temp['B6'].value = room['B']
    except:
        sheet_obj_temp['B6'].value = 0
    try:
        sheet_obj_temp['B7'].value = room['D']
    except:
        sheet_obj_temp['B7'].value = 0
    try:
        sheet_obj_temp['B8'].value = room['E']
    except:
        sheet_obj_temp['B8'].value = 0
    try:
        sheet_obj_temp['B9'].value = room['G']
    except:
        sheet_obj_temp['B9'].value = 0
    try:
        sheet_obj_temp['B10'].value = room['K']
    except:
        sheet_obj_temp['B10'].value = 0
    try:    
        sheet_obj_temp['B11'].value = room['L']
    except:
        sheet_obj_temp['B11'].value = 0
    try:
        sheet_obj_temp['B12'].value = room['M']
    except:
        sheet_obj_temp['B12'].value = 0
    try:
        sheet_obj_temp['B13'].value = room['N']
    except:
        sheet_obj_temp['B13'].value = 0
    try:
        sheet_obj_temp['B14'].value = room['P']
    except:
        sheet_obj_temp['B14'].value = 0
    try:
        sheet_obj_temp['B15'].value = room['R']
    except:
        sheet_obj_temp['B15'].value = 0
    try:
        sheet_obj_temp['B16'].value = room['S']
    except:
        sheet_obj_temp['B16'].value = 0
    try:
        sheet_obj_temp['B17'].value = room['U']
    except:
        sheet_obj_temp['B17'].value = 0
    try:
        sheet_obj_temp['B18'].value = room['V']
    except:
        sheet_obj_temp['B18'].value = 0
    try:
        sheet_obj_temp['B19'].value = room['Y']
    except:
        sheet_obj_temp['B19'].value = 0
    try:
        sheet_obj_temp['B20'].value = room['W']
    except:
        sheet_obj_temp['B20'].value = 0
    try:
        sheet_obj_temp['B21'].value = room['Z']
    except:
        sheet_obj_temp['B21'].value = 0
    os.makedirs(f"{output_path}\\DRR_excel",exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\opera_drr.xlsx")

def rev_val(revenue,output_path):
    
    wb_obj_temp = load_workbook(f"{output_path}\\DRR_excel\\opera_drr.xlsx")
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column
    try:
        sheet_obj_temp['C5'].value = str(revenue['A'])
    except:
        sheet_obj_temp['C5'].value = 0
    try:
        sheet_obj_temp['C6'].value = str(revenue['B'])
    except:
        sheet_obj_temp['C6'].value = 0
    try:
        sheet_obj_temp['C7'].value = str(revenue['D'])
    except:
        sheet_obj_temp['C7'].value = 0
    try:
        sheet_obj_temp['C8'].value = str(revenue['E'])
    except:
        sheet_obj_temp['C8'].value = 0
    try:
        sheet_obj_temp['C9'].value = str(revenue['G'])
    except:
        sheet_obj_temp['C9'].value = 0
    try:
        sheet_obj_temp['C10'].value = str(revenue['K'])
    except:
        sheet_obj_temp['C10'].value = 0
    try:    
        sheet_obj_temp['C11'].value = str(revenue['L'])
    except:
        sheet_obj_temp['C11'].value = 0
    try:
        sheet_obj_temp['C12'].value = str(revenue['M'])
    except:
        sheet_obj_temp['C12'].value = 0
    try:
        sheet_obj_temp['C13'].value = str(revenue['N'])
    except:
        sheet_obj_temp['C13'].value = 0
    try:
        sheet_obj_temp['C14'].value = str(revenue['P'])
    except:
        sheet_obj_temp['C14'].value = 0
    try:
        sheet_obj_temp['C15'].value = str(revenue['R'])
    except:
        sheet_obj_temp['C15'].value = 0
    try:
        sheet_obj_temp['C16'].value = str(revenue['S'])
    except:
        sheet_obj_temp['C16'].value = 0
    try:
        sheet_obj_temp['C17'].value = str(revenue['U'])
    except:
        sheet_obj_temp['C17'].value = 0
    try:
        sheet_obj_temp['C18'].value = str(revenue['V'])
    except:
        sheet_obj_temp['C18'].value = 0
    try:
        sheet_obj_temp['C19'].value = str(revenue['Y'])
    except:
        sheet_obj_temp['C19'].value = 0
    try:
        sheet_obj_temp['C20'].value = str(revenue['W'])
    except:
        sheet_obj_temp['C20'].value = 0
    try:
        sheet_obj_temp['C21'].value = str(revenue['Z'])
    except:
        sheet_obj_temp['C21'].value = 0

    wb_obj_temp.save(f"{output_path}\\DRR_excel\\opera_drr.xlsx")

def manager_report(input_path,output_path):
    
    wb_obj = load_workbook(f'{input_path}\\manager_report.xlsx')
    sheet_obj = wb_obj.active
    rows = sheet_obj.max_row
    cols = sheet_obj.max_column

    key = []
    value = []
    for i in range(1,rows+1):
        cell_obj = sheet_obj.cell(row=i, column=2)
        keys = cell_obj.value
        key.append(keys)
    for i in range(1,rows+1):
        cell_obj = sheet_obj.cell(row=i, column=3)
        val = cell_obj.value
        value.append(val)

    manager_repot = dict(zip(key, value))

    path_template = f"{output_path}\\DRR_excel\\opera_drr.xlsx"
    wb_obj_temp = load_workbook(path_template)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    try:
        sheet_obj_temp['H1'].value = int(manager_repot['Total Rooms in Hotel'].split(' ')[0])
    except:
        sheet_obj_temp['H1'].value = 0

    try:
        sheet_obj_temp['J6'].value = int(manager_repot['Rooms Occupied'].split(' ')[0])
    except:
        sheet_obj_temp['J6'].value = 0

    try:
        sheet_obj_temp['J7'].value = int(manager_repot['Available Rooms minus OOO Rooms'].split(' ')[0])
    except:
        sheet_obj_temp['J7'].value = 0

    try:
        sheet_obj_temp['J8'].value = 0
    except:
        sheet_obj_temp['J8'].value = 0

    try:
        sheet_obj_temp['J9'].value = int(manager_repot['Out of Order Rooms'].split(' ')[0])
    except:
        sheet_obj_temp['J9'].value = 0

    wb_obj_temp.save(f"{output_path}\\DRR_excel\\opera_drr.xlsx")


def marketformat(file_path):
    wb_obj = load_workbook(file_path)
    sheet_obj = wb_obj.active
    rows = sheet_obj.max_row
    cols = sheet_obj.max_column
    dr=0
    dc=0
    for i in range(1,rows+1):
        for j in range(1,cols+1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            if cell_obj.value is None:
                continue
            cell_data=str(cell_obj.value)
            if "description" in cell_data.lower():
                dr=i
                dc=j
    return (dr,dc-1)



def convert_tabula_to_drr(path):
    
    tabula_path=f'{path}\\tabula_output\\market_code_stats_report'
    print(path)
    key_lock=0
    key = []
    val_room = []
    val_revenue = []
    for xlfile in os.listdir(tabula_path):
        # print(xlfile)
        wb_obj = load_workbook(f'{path}\\tabula_output\\market_code_stats_report\\{xlfile}')
        sheet_obj = wb_obj.active
        rows = sheet_obj.max_row
        cols = sheet_obj.max_column
        
        (start_row,start_col)=marketformat(f"{tabula_path}\\{xlfile}")
        # print(start_row,',',start_col)
        i=start_row+1
        while i<=rows:
            cell_obj = sheet_obj.cell(row=i, column=start_col+1)
            cell_data=str(cell_obj.value)
            if key_lock==0 :
                cell_obj = sheet_obj.cell(row=i, column=start_col+1)
                cell_data=str(cell_obj.value)
                keys=cell_data
                i+=1
                key_lock=1
            while i<=rows and key_lock==1:
                cell_obj = sheet_obj.cell(row=i, column=start_col)
                cell_data=str(cell_obj.value) 
                if "group total" in cell_data.lower():
                    room_obj=sheet_obj.cell(row=i,column=start_col+2)
                    room_value=room_obj.value
                    revenue_obj=sheet_obj.cell(row=i,column=start_col+3)
                    revenue_value=revenue_obj.value
                    key.append(keys.split(' ')[0])
                    val_room.append(room_value)
                    val_revenue.append(revenue_value)
                    print  (keys,'\n\tRooms -',room_value,'\n\tRevenue - ',revenue_value) 
                    key_lock=0
                i+=1    
       

    room = dict(zip(key, val_room))
    revenue = dict(zip(key,val_revenue))

    room_val(room,path)
    rev_val(revenue,path)
    manager_report(f'{path}\\tabula_output\\manager_report',path)


#  ADD fossee code 

def f1_read_text_files(fosse_folder_path):
    file_loc = config["file_path"]["loc"]
    y = datetime.datetime.now()
    mon = y.strftime("%B")
    t_date = date.today()
    yesterday = t_date-datetime.timedelta(days=1)
    previous_date = yesterday.strftime("%d")
    # for folder in os.listdir(f'{file_loc}'):
    folder = fosse_folder_path

    if properties_mapping[f'{folder}'] == "Fossee_type1":
        
        path= f'{file_loc}\\{folder}\\July\\13'
        for file_name in os.listdir(path):
            if 'stats' in file_name.lower():
                f1_stats(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'segment' in file_name.lower():
                f1_segmentation(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'rev' in file_name.lower() and '.txt' in file_name.lower():
                f1_revenue(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'ledger' in file_name.lower():
                f1_ledger(f'{path}\\{file_name}',f'{file_loc}\\{folder}\\DRR-Template.xlsx',path)

    if properties_mapping[f'{folder}'] == "Fossee_type2":
        
        path= f'{file_loc}\\{folder}\\July\\13'
        for file_name in os.listdir(path):
            if 'stat' in file_name.lower():
                f1_stats(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'seg' in file_name.lower():
                f1_segmentation(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'rev' in file_name.lower() and '.txt' in file_name.lower():
                f2_revenue(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'ledger' in file_name.lower():
                f1_ledger(f'{path}\\{file_name}',f'{file_loc}\\{folder}\\DRR-Template.xlsx',path)



    if properties_mapping[f'{folder}'] == "Fossee_type3":
        
        path= f'{file_loc}\\{folder}\\July\\13'
        for file_name in os.listdir(path):
            if 'stat' in file_name.lower():
                f1_stats(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'seg' in file_name.lower():
                f3_segmentation(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'rev' in file_name.lower() and '.txt' in file_name.lower():
                f3_revenue(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

            if 'ledger' in file_name.lower():
                f3_ledger(f'{path}\\{file_name}',f'{file_loc}\\{folder}\\DRR-Template.xlsx',path)



def f1_revenue(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        if "   ROOMS   " in lines:
            x=re.search("   ROOMS   ",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B17'].value= str(res)
            


        if "PE  Pet Charge" in lines:
            x=re.search("PE  Pet Charge",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B20'].value= str(res)
            
        if "GS  Gift Shop" in lines:
            x=re.search("GS  Gift Shop",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B21'].value= str(res)
            
        if "VC  Vending Commissions" in lines:
            x=re.search("VC  Vending Commissions",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B22'].value= str(res)
            
        if "MH  Misc" in lines:
            x=re.search("MH  Misc",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B23'].value= str(res)
            
        if "LONG DISTANCE PHONE" in lines:
            x=re.search("LONG DISTANCE PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B24'].value= str(res)
            
        
        
        if "State Occupancy Tax" in lines:
            x=re.search("State Occupancy Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B33'].value= str(res)
            

        if "T3  City Tax" in lines:
            x=re.search("City Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B34'].value= str(res)
            

        if "T9  Sales Tax" in lines:
            x=re.search("Sales Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B35'].value= str(res)
            

        if "CA  Cash" in lines:
            x=re.search("CA  Cash",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B38'].value= str(res)
            

        if "CK  Check" in lines:
            x=re.search("CK  Check",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B39'].value= str(res)
            

        if "AX  American Express" in lines:
            x=re.search("AX  American Express",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B40'].value= str(res)
            

        if "MC  Master Card" in lines:
            x=re.search("MC  Master Card",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B41'].value= str(res)
            

        if "VI  Visa" in lines:
            x=re.search("VI  Visa",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B42'].value= str(res)
            

        if "DS  Discover" in lines:
            x=re.search("DS  Discover",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B43'].value= str(res)
            

        if "WT  Wire Transfer" in lines:
            x=re.search("WT  Wire Transfer",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B44'].value= str(res)
            

        if "MBV Redemption" in lines:
            x=re.search("MBV Redemption",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B44'].value= str(res)
            


        

        


    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")

        



def f1_ledger(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        
        if "A D V A N C E" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B63'].value= str(res)
            

            while not "NEW DEPOSITS BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("NEW DEPOSITS BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C63'].value= str(res)
            
            
        
        if "G U E S T    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B61'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C61'].value= str(res)
            
        
        if "C I T Y    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B62'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C62'].value= str(res)
            
        
        line_number+=1

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")


def f1_segmentation(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()

    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    for lines in content.split('\n'):
        
        if 'BTBP' in lines.upper() or 'TSGR' in lines.upper(): 
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B5'].value= int(sheet_obj_temp['B5'].value) + int(res[0])
            sheet_obj_temp['C5'].value= float(sheet_obj_temp['C5'].value) + float(res[1])

        if 'REG' in lines.upper():
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B6'].value= int(sheet_obj_temp['B6'].value) + int(res[0])
            sheet_obj_temp['C6'].value= float(sheet_obj_temp['C6'].value) + float(res[1])

        if 'GOV' in lines.upper():
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B7'].value= int(sheet_obj_temp['B7'].value) + int(res[0])
            sheet_obj_temp['C7'].value= float(sheet_obj_temp['C7'].value) + float(res[1])

        if 'AABS' in lines.upper() or 'QAAD' in lines.upper() or 'APND' in lines.upper() or 'RMO' in lines.upper() or 'MRY' in lines.upper():
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B9'].value= int(sheet_obj_temp['B9'].value) + int(res[0])
            sheet_obj_temp['C9'].value= float(sheet_obj_temp['C9'].value) + float(res[1])

        if 'KXPK' in lines.upper() or 'XXPA' in lines.upper() or 'XXPB' in lines.upper() :
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B10'].value= int(sheet_obj_temp['B10'].value) + int(res[0])
            sheet_obj_temp['C10'].value= float(sheet_obj_temp['C10'].value) + float(res[1])

        if 'LTSA' in lines.upper() or 'LTSB' in lines.upper() :
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B11'].value= int(sheet_obj_temp['B11'].value) + int(res[0])
            sheet_obj_temp['C11'].value= float(sheet_obj_temp['C11'].value) + float(res[1])
        
        if re.findall("^17", lines) or re.findall("^18", lines) or re.findall("^24", lines) or re.findall("^28", lines) or  re.findall("^36", lines) or re.findall("^43", lines) or re.findall("^44", lines):
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B8'].value= int(sheet_obj_temp['B8'].value) + int(res[0])
            sheet_obj_temp['C8'].value= float(sheet_obj_temp['C8'].value) + float(res[1])

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")

        
        




def f1_stats(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()

    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    for lines in content.split('\n'):
        
        if 'ROOMS SOLD' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L3'].value= int(res)

        if 'ROOMS VACANT' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L4'].value= int(res) 

        if 'OUT OF ORDER' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L5'].value= int(res)   

        if 'COMPLIMENTARY ROOMS' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L6'].value= int(res)  

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")
   

def f2_revenue(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        if "   ROOMS   " in lines:
            x=re.search("   ROOMS   ",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B17'].value= str(res)
            


        if "PE  Pet Charge" in lines:
            x=re.search("PE  Pet Charge",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B20'].value= str(res)
            
        if "GS  Gift Shop" in lines:
            x=re.search("GS  Gift Shop",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B21'].value= str(res)
            
        if "VC  Vending Commissions" in lines:
            x=re.search("VC  Vending Commissions",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B22'].value= str(res)
            
        if "MH  Misc" in lines:
            x=re.search("MH  Misc",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B23'].value= str(res)
            
        if "LONG DISTANCE PHONE" in lines:
            x=re.search("LONG DISTANCE PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B24'].value= str(res)
            
        
        
        if "State Occupancy Tax" in lines:
            x=re.search("State Occupancy Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B27'].value= str(res)
            

        if "T3  City Tax" in lines:
            x=re.search("City Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B28'].value= str(res)
            

        if "T9  Sales Tax" in lines:
            x=re.search("Sales Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B29'].value= str(res)
            

        if "CA  Cash" in lines:
            x=re.search("CA  Cash",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B32'].value= float(res)
            

        if "CK  Check" in lines:
            x=re.search("CK  Check",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B32'].value= float(sheet_obj_temp['B32'].value) + float(res)
            

        if "AX  American Express" in lines:
            x=re.search("AX  American Express",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B33'].value= str(res)
            

        if "MC  Master Card" in lines:
            x=re.search("MC  Master Card",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B34'].value= str(res)
            

        if "VI  Visa" in lines:
            x=re.search("VI  Visa",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B35'].value= str(res)
            

        if "DS  Discover" in lines:
            x=re.search("DS  Discover",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B36'].value= str(res)
            

        if "WT  Wire Transfer" in lines:
            x=re.search("WT  Wire Transfer",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B37'].value= str(res)
            

        if "MBV Redemption" in lines:
            x=re.search("MBV Redemption",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B39'].value= str(res)


def f2_ledger(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        
        if "A D V A N C E" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B56'].value= str(res)
            

            while not "NEW DEPOSITS BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("NEW DEPOSITS BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C56'].value= str(res)
            
            
        
        if "G U E S T    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B54'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C54'].value= str(res)
            
        
        if "C I T Y    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B55'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C55'].value= str(res)
            
        
        line_number+=1

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")


def f3_revenue(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        if "   ROOMS   " in lines:
            x=re.search("   ROOMS   ",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B18'].value= str(res)
            print (res)


        if "PE  Pet Charge" in lines:
            x=re.search("PE  Pet Charge",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B22'].value= str(res)
            print (res)


        if "GS  Gift Shop" in lines:
            x=re.search("GS  Gift Shop",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B23'].value= str(res)
            print (res)

        if "VC  Vending Commissions" in lines:
            x=re.search("VC  Vending Commissions",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B24'].value= str(res)
            print (res)


        if "WA  Washer/Dryer Income" in lines:
            x=re.search("WA  Washer/Dryer Income",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B25'].value= str(res)
            print (res)

        if "FUNCTION ROOMS" in lines:
            x=re.search("FUNCTION ROOMS",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B26'].value= str(res)
            print (res)

        if "MH  Misc" in lines:
            x=re.search("MH  Misc",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B27'].value= str(res)
            print (res)


        if "LOCAL PHONE" in lines:
            x=re.search("LOCAL PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B31'].value= str(res)
            print (res)


        if "LONG DISTANCE PHONE" in lines:
            x=re.search("LONG DISTANCE PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B32'].value= str(res)
            print (res)
        
        
        if "State Occupancy Tax" in lines:
            x=re.search("State Occupancy Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B36'].value= str(res)
            print (res)

        if "T3  City Tax" in lines:
            x=re.search("T3  City Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B37'].value= str(res)
            print (res)


        if "T4  County Tax" in lines:
            x=re.search("T4  County Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B38'].value= str(res)
            print (res)

        if "T5  Convention and Touris" in lines:
            x=re.search("T5  Convention and Touris",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B39'].value= str(res)
            print (res)

        if "T9  Sales Tax" in lines:
            x=re.search("T9  Sales Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B40'].value= str(res)
            print (res)

        if "CA  Cash" in lines:
            x=re.search("CA  Cash",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B45'].value= str(res)
            print (res)

        if "CK  Check" in lines:
            x=re.search("CK  Check",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B46'].value= str(res)
            print (res)

        if "AX  American Express" in lines:
            x=re.search("AX  American Express",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B47'].value= str(res)
            print (res)

        if "MC  Master Card" in lines:
            x=re.search("MC  Master Card",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B48'].value= str(res)
            print (res)

        if "VI  Visa" in lines:
            x=re.search("VI  Visa",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B49'].value= str(res)
            print (res)

        if "DS  Discover" in lines:
            x=re.search("DS  Discover",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B50'].value= str(res)
            print (res)

        if "WT  Wire Transfer" in lines:
            x=re.search("WT  Wire Transfer",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B51'].value= str(res)
            print (res)


        

        


    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")

        



def f3_ledger(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        
        if "A D V A N C E" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B71'].value= str(res)
            

            while not "NEW DEPOSITS BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("NEW DEPOSITS BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C71'].value= str(res)
            
            
        
        if "G U E S T    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B69'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C69'].value= str(res)
            
        
        if "C I T Y    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B70'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C70'].value= str(res)
            
        
        line_number+=1

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")




# ADD onq code



# __main__
for file_name in properties_mapping:
    if properties_mapping[file_name] == "opera":
        drr_excel(file_name)

    if properties_mapping[file_name] == "Fossee_type1":
        f1_read_text_files(file_name)

    # if properties_mapping[file_name] == ""