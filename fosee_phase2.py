import tabula
import os
import re
import openpyxl
import csv
from openpyxl import load_workbook
from datetime import date
import pandas as pd
import glob
from openpyxl import workbook
t_date = date.today()
import pdfplumber
import sys
import logging
import datetime
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

logging.basicConfig(filename='Drr_generator.log',
                    format='%(asctime)s %(message)s',
                    filemode='a',
                    level = logging.DEBUG)


import configparser
config = configparser.ConfigParser()
config.read('DRR.config.ini')

properties_mapping ={"Night Audit- Towneplace Suites Houston NW":"Fossee_type1","Night Audit- Towneplace Plano":"Fossee_type1"}

def f1_read_text_files():
    file_loc = config["file_path"]["loc"]
    y = datetime.datetime.now()
    mon = y.strftime("%B")
    t_date = date.today()
    yesterday = t_date-datetime.timedelta(days=1)
    previous_date = yesterday.strftime("%d")
    for folder in os.listdir(f'{file_loc}'):

        if properties_mapping[f'{folder}'] == "Fossee_type1":
            
            path= f'{file_loc}\\{folder}\\July\\13'
            for file_name in os.listdir(path):
                if 'stats' in file_name.lower():
                    f1_stats(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'segment' in file_name.lower():
                    f1_segmentation(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'rev' in file_name.lower() and '.txt' in file_name.lower():
                    f1_revenue(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'ledger' in file_name.lower():
                    f1_ledger(f'{path}\\{file_name}',f'{file_loc}\\{folder}\\DRR-Template.xlsx',path)

        if properties_mapping[f'{folder}'] == "Fossee_type2":
            
            path= f'{file_loc}\\{folder}\\July\\13'
            for file_name in os.listdir(path):
                if 'stat' in file_name.lower():
                    f1_stats(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'seg' in file_name.lower():
                    f1_segmentation(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'rev' in file_name.lower() and '.txt' in file_name.lower():
                    f2_revenue(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'ledger' in file_name.lower():
                    f1_ledger(f'{path}\\{file_name}',f'{file_loc}\\{folder}\\DRR-Template.xlsx',path)



        if properties_mapping[f'{folder}'] == "Fossee_type3":
            
            path= f'{file_loc}\\{folder}\\July\\13'
            for file_name in os.listdir(path):
                if 'stat' in file_name.lower():
                    f1_stats(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'seg' in file_name.lower():
                    f3_segmentation(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'rev' in file_name.lower() and '.txt' in file_name.lower():
                    f3_revenue(f'{path}\\{file_name}',f'{path}\\DRR_excel\\fossee_drr.xlsx',path)

                if 'ledger' in file_name.lower():
                    f3_ledger(f'{path}\\{file_name}',f'{file_loc}\\{folder}\\DRR-Template.xlsx',path)



def f1_revenue(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        if "   ROOMS   " in lines:
            x=re.search("   ROOMS   ",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B17'].value= str(res)
            


        if "PE  Pet Charge" in lines:
            x=re.search("PE  Pet Charge",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B20'].value= str(res)
            
        if "GS  Gift Shop" in lines:
            x=re.search("GS  Gift Shop",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B21'].value= str(res)
            
        if "VC  Vending Commissions" in lines:
            x=re.search("VC  Vending Commissions",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B22'].value= str(res)
            
        if "MH  Misc" in lines:
            x=re.search("MH  Misc",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B23'].value= str(res)
            
        if "LONG DISTANCE PHONE" in lines:
            x=re.search("LONG DISTANCE PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B24'].value= str(res)
            
        
        
        if "State Occupancy Tax" in lines:
            x=re.search("State Occupancy Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B33'].value= str(res)
            

        if "T3  City Tax" in lines:
            x=re.search("City Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B34'].value= str(res)
            

        if "T9  Sales Tax" in lines:
            x=re.search("Sales Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B35'].value= str(res)
            

        if "CA  Cash" in lines:
            x=re.search("CA  Cash",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B38'].value= str(res)
            

        if "CK  Check" in lines:
            x=re.search("CK  Check",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B39'].value= str(res)
            

        if "AX  American Express" in lines:
            x=re.search("AX  American Express",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B40'].value= str(res)
            

        if "MC  Master Card" in lines:
            x=re.search("MC  Master Card",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B41'].value= str(res)
            

        if "VI  Visa" in lines:
            x=re.search("VI  Visa",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B42'].value= str(res)
            

        if "DS  Discover" in lines:
            x=re.search("DS  Discover",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B43'].value= str(res)
            

        if "WT  Wire Transfer" in lines:
            x=re.search("WT  Wire Transfer",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B44'].value= str(res)
            

        if "MBV Redemption" in lines:
            x=re.search("MBV Redemption",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B44'].value= str(res)
            


        

        


    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")

        



def f1_ledger(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        
        if "A D V A N C E" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B63'].value= str(res)
            

            while not "NEW DEPOSITS BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("NEW DEPOSITS BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C63'].value= str(res)
            
            
        
        if "G U E S T    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B61'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C61'].value= str(res)
            
        
        if "C I T Y    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B62'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C62'].value= str(res)
            
        
        line_number+=1

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")


def f1_segmentation(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()

    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    for lines in content.split('\n'):
        
        if 'BTBP' in lines.upper() or 'TSGR' in lines.upper(): 
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B5'].value= int(sheet_obj_temp['B5'].value) + int(res[0])
            sheet_obj_temp['C5'].value= float(sheet_obj_temp['C5'].value) + float(res[1])

        if 'REG' in lines.upper():
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B6'].value= int(sheet_obj_temp['B6'].value) + int(res[0])
            sheet_obj_temp['C6'].value= float(sheet_obj_temp['C6'].value) + float(res[1])

        if 'GOV' in lines.upper():
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B7'].value= int(sheet_obj_temp['B7'].value) + int(res[0])
            sheet_obj_temp['C7'].value= float(sheet_obj_temp['C7'].value) + float(res[1])

        if 'AABS' in lines.upper() or 'QAAD' in lines.upper() or 'APND' in lines.upper() or 'RMO' in lines.upper() or 'MRY' in lines.upper():
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B9'].value= int(sheet_obj_temp['B9'].value) + int(res[0])
            sheet_obj_temp['C9'].value= float(sheet_obj_temp['C9'].value) + float(res[1])

        if 'KXPK' in lines.upper() or 'XXPA' in lines.upper() or 'XXPB' in lines.upper() :
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B10'].value= int(sheet_obj_temp['B10'].value) + int(res[0])
            sheet_obj_temp['C10'].value= float(sheet_obj_temp['C10'].value) + float(res[1])

        if 'LTSA' in lines.upper() or 'LTSB' in lines.upper() :
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B11'].value= int(sheet_obj_temp['B11'].value) + int(res[0])
            sheet_obj_temp['C11'].value= float(sheet_obj_temp['C11'].value) + float(res[1])
        
        if re.findall("^17", lines) or re.findall("^18", lines) or re.findall("^24", lines) or re.findall("^28", lines) or  re.findall("^36", lines) or re.findall("^43", lines) or re.findall("^44", lines):
            res = re.findall('[\d\,\.\d]+\W', lines)
            sheet_obj_temp['B8'].value= int(sheet_obj_temp['B8'].value) + int(res[0])
            sheet_obj_temp['C8'].value= float(sheet_obj_temp['C8'].value) + float(res[1])

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")

        
        




def f1_stats(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()

    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    for lines in content.split('\n'):
        
        if 'ROOMS SOLD' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L3'].value= int(res)

        if 'ROOMS VACANT' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L4'].value= int(res) 

        if 'OUT OF ORDER' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L5'].value= int(res)   

        if 'COMPLIMENTARY ROOMS' in lines.upper():
            res = re.findall('[\d\,\.\d]+', lines)[0]
            sheet_obj_temp['L6'].value= int(res)  

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")
   

def f2_revenue(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        if "   ROOMS   " in lines:
            x=re.search("   ROOMS   ",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B17'].value= str(res)
            


        if "PE  Pet Charge" in lines:
            x=re.search("PE  Pet Charge",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B20'].value= str(res)
            
        if "GS  Gift Shop" in lines:
            x=re.search("GS  Gift Shop",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B21'].value= str(res)
            
        if "VC  Vending Commissions" in lines:
            x=re.search("VC  Vending Commissions",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B22'].value= str(res)
            
        if "MH  Misc" in lines:
            x=re.search("MH  Misc",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B23'].value= str(res)
            
        if "LONG DISTANCE PHONE" in lines:
            x=re.search("LONG DISTANCE PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B24'].value= str(res)
            
        
        
        if "State Occupancy Tax" in lines:
            x=re.search("State Occupancy Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B27'].value= str(res)
            

        if "T3  City Tax" in lines:
            x=re.search("City Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B28'].value= str(res)
            

        if "T9  Sales Tax" in lines:
            x=re.search("Sales Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B29'].value= str(res)
            

        if "CA  Cash" in lines:
            x=re.search("CA  Cash",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B32'].value= float(res)
            

        if "CK  Check" in lines:
            x=re.search("CK  Check",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B32'].value= float(sheet_obj_temp['B32'].value) + float(res)
            

        if "AX  American Express" in lines:
            x=re.search("AX  American Express",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B33'].value= str(res)
            

        if "MC  Master Card" in lines:
            x=re.search("MC  Master Card",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B34'].value= str(res)
            

        if "VI  Visa" in lines:
            x=re.search("VI  Visa",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B35'].value= str(res)
            

        if "DS  Discover" in lines:
            x=re.search("DS  Discover",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B36'].value= str(res)
            

        if "WT  Wire Transfer" in lines:
            x=re.search("WT  Wire Transfer",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B37'].value= str(res)
            

        if "MBV Redemption" in lines:
            x=re.search("MBV Redemption",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B39'].value= str(res)


def f2_ledger(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        
        if "A D V A N C E" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B56'].value= str(res)
            

            while not "NEW DEPOSITS BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("NEW DEPOSITS BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C56'].value= str(res)
            
            
        
        if "G U E S T    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B54'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C54'].value= str(res)
            
        
        if "C I T Y    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B55'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C55'].value= str(res)
            
        
        line_number+=1

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")


def f3_revenue(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        if "   ROOMS   " in lines:
            x=re.search("   ROOMS   ",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B18'].value= str(res)
            print (res)


        if "PE  Pet Charge" in lines:
            x=re.search("PE  Pet Charge",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B22'].value= str(res)
            print (res)


        if "GS  Gift Shop" in lines:
            x=re.search("GS  Gift Shop",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B23'].value= str(res)
            print (res)

        if "VC  Vending Commissions" in lines:
            x=re.search("VC  Vending Commissions",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B24'].value= str(res)
            print (res)


        if "WA  Washer/Dryer Income" in lines:
            x=re.search("WA  Washer/Dryer Income",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B25'].value= str(res)
            print (res)

        if "FUNCTION ROOMS" in lines:
            x=re.search("FUNCTION ROOMS",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B26'].value= str(res)
            print (res)

        if "MH  Misc" in lines:
            x=re.search("MH  Misc",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B27'].value= str(res)
            print (res)


        if "LOCAL PHONE" in lines:
            x=re.search("LOCAL PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B31'].value= str(res)
            print (res)


        if "LONG DISTANCE PHONE" in lines:
            x=re.search("LONG DISTANCE PHONE",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B32'].value= str(res)
            print (res)
        
        
        if "State Occupancy Tax" in lines:
            x=re.search("State Occupancy Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B36'].value= str(res)
            print (res)

        if "T3  City Tax" in lines:
            x=re.search("T3  City Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B37'].value= str(res)
            print (res)


        if "T4  County Tax" in lines:
            x=re.search("T4  County Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B38'].value= str(res)
            print (res)

        if "T5  Convention and Touris" in lines:
            x=re.search("T5  Convention and Touris",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B39'].value= str(res)
            print (res)

        if "T9  Sales Tax" in lines:
            x=re.search("T9  Sales Tax",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B40'].value= str(res)
            print (res)

        if "CA  Cash" in lines:
            x=re.search("CA  Cash",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B45'].value= str(res)
            print (res)

        if "CK  Check" in lines:
            x=re.search("CK  Check",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B46'].value= str(res)
            print (res)

        if "AX  American Express" in lines:
            x=re.search("AX  American Express",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B47'].value= str(res)
            print (res)

        if "MC  Master Card" in lines:
            x=re.search("MC  Master Card",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B48'].value= str(res)
            print (res)

        if "VI  Visa" in lines:
            x=re.search("VI  Visa",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B49'].value= str(res)
            print (res)

        if "DS  Discover" in lines:
            x=re.search("DS  Discover",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B50'].value= str(res)
            print (res)

        if "WT  Wire Transfer" in lines:
            x=re.search("WT  Wire Transfer",lines).start()
            res = str(re.findall('[\d\,\.\d]+\W', lines[x:-1])[1]) 
            sheet_obj_temp['B51'].value= str(res)
            print (res)


        

        


    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")

        



def f3_ledger(text_file,template_file,output_path):
    file_var=open(text_file,'r')
    content= file_var.read()


    wb_obj_temp = load_workbook(template_file)
    sheet_obj_temp = wb_obj_temp.active
    rows_temp = sheet_obj_temp.max_row
    cols_temp = sheet_obj_temp.max_column

    list_of_lines=content.split('\n')
    line_number=0
    for lines in list_of_lines:

        
        if "A D V A N C E" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B71'].value= str(res)
            

            while not "NEW DEPOSITS BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("NEW DEPOSITS BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C71'].value= str(res)
            
            
        
        if "G U E S T    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B69'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C69'].value= str(res)
            
        
        if "C I T Y    L E D G E R" in lines.upper():
            
            i=line_number+1
            while not "BALANCE" in list_of_lines[i].upper():
                i+=1
            x=re.search("BALANCE",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['B70'].value= str(res)
            

            while not "CURRENT" in list_of_lines[i].upper():
                i+=1
            x=re.search("CURRENT",list_of_lines[i].upper()).start() 
            list_of_lines[i]+="   "
            res = re.findall('[\d\,\d\.\d]+\W', list_of_lines[i][x:-1])[0]
            sheet_obj_temp['C70'].value= str(res)
            
        
        line_number+=1

    os.makedirs(f"{output_path}\\DRR_excel", exist_ok=True)
    wb_obj_temp.save(f"{output_path}\\DRR_excel\\fossee_drr.xlsx")




f1_read_text_files()